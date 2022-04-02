import openpyxl
import pandas
import pandas as pd
import xlsxwriter
# читаем excel-файл (в скобках указать путь к файлу)
wb = openpyxl.load_workbook('C:\\Users\Admin\selenium_course\\fortest.xlsx')
# получаем активный лист
sheet = wb.active
# Переменные для определения в каких ячейках находятся нужные нам поля
Created_time = 0
Click_id = 0
Release_name = 0
Status = 0
rel_Status = 0
big = ''
# Определяем , чем и куда будем записывать итоговый файл(указать путь файла)
writer = pd.ExcelWriter('C:\\Users\Admin\selenium_course\\itogitog.xlsx', engine='xlsxwriter')
# Создаём листы для экселя
df = pandas.DataFrame({"Id клика": [],
                       "Номер": [],
                       "Время": [],
                       "Статус": [],
                       "Статус": []
                       })
trouble_df = pandas.DataFrame({"Номер": []
                               })

NotRec_Rec_Itog = pandas.DataFrame({"Номер": [],
                       "Начальный статус": [],
                       "Время начальное": [],
                       "Финальный статус": [],
                       "Время финальное": []
                       })
Rec_Rec_Itog = pandas.DataFrame({"Номер": [],
                       "Начальный статус": [],
                       "Время начальное": [],
                       "Финальный статус": [],
                       "Время финальное": []
                       })
Rec_NotRec_Itog = pandas.DataFrame({"Номер": [],
                       "Начальный статус": [],
                       "Время начальное": [],
                       "Финальный статус": [],
                       "Время финальное": []
                       })
NotRec_NotRec_Itog = pandas.DataFrame({"Номер": [],
                       "Начальный статус": [],
                       "Время начальное": [],
                       "Финальный статус": [],
                       "Время финальное": []
                       })

# Определяем количество строк в файле, потом проходимся циклом для определения названия полей
header_cells_generator = sheet.iter_rows(max_row=1)
for header_cells_tuple in header_cells_generator:
    for i in range(len(header_cells_tuple)):
        # разкомитить строку ниже, если нужно посмотреть названия полей
        # print(header_cells_tuple[i].value, end="  ")
        if header_cells_tuple[i].value == 'CREATED_AT':
            Created_time = i + 1
        if header_cells_tuple[i].value == 'ID':
            Click_id = i + 1
        if header_cells_tuple[i].value == 'KEY':
            Release_name = i + 1
        if header_cells_tuple[i].value == 'RESOLUTION':
            Status = i + 1
        if header_cells_tuple[i].value == 'STATUS':
            rel_Status = i + 1

# Определяем массивы для сохранения в них таблицы
release_key = []
created_time = []
ckick_id = []
status_ift = []
status_of_release = []
key_key = []
check = []
# Скачиваем в массив всю таблицу экселя и запихиваем в массивы
for x in range(2, sheet.max_row + 1):
    a = sheet.cell(row=x, column=Created_time).value
    b = sheet.cell(row=x, column=Click_id).value
    c = sheet.cell(row=x, column=Release_name).value
    d = sheet.cell(row=x, column=Status).value
    e = sheet.cell(row=x, column=rel_Status).value
    # добавляем номер в общий список
    release_key.append(c)
    created_time.append(a)
    ckick_id.append(b)
    status_ift.append(d)
    status_of_release.append(e)
    # Добавляем номер строки в список
    key_key.append(x)
# Получаем уникальные названия , без дублей
double1 = set(release_key)
# Определяем итоговые массивы , куда будем заносить названия
NotRec_NotRec = []
NotRec_Rec = []
Rec_Rec = []
Rec_NotRec = []
trouble = []

# основной метод Берём первое уникальное название из списка double1 и проходимся по всему массиву release_key
# ища одинаковые названия. У этого вытягиваем дату, определяем максимальную и минимальную дату и смотрим на
# статусы в эти даты, сортируя в нужные массивы. В конце формируем строки для датафреймов с max и min значениями

def relis1(arg):
    global NotRec_Rec_Itog
    global Rec_Rec_Itog
    global Rec_NotRec_Itog
    global NotRec_NotRec_Itog
    sravn = []
    for k in range(len(release_key)):
        if release_key[k] == arg:
                # print(ckick_id[k], release_key[k], created_time[k])
            sravn.append(created_time[k])
    for k in range(len(release_key)):
        if created_time[k] == min(sravn):
            minim = k
        if created_time[k] == max(sravn):
            maxi = k

        # print(ckick_id[minim], release_key[minim], created_time[minim], status_ift[minim],status_of_release[minim])
        # print(ckick_id[maxi], release_key[maxi], created_time[maxi], status_ift[maxi],status_of_release[maxi])

    if status_ift[maxi] == 'Рекомендован' and status_ift[minim] == 'Не рекомендован':
        NotRec_Rec.append(release_key[maxi])
        NotRec_Rec_row = {"Номер": release_key[minim], "Начальный статус": status_of_release[minim],
                        "Время начальное": created_time[minim], "Финальный статус": status_of_release[maxi],
                        "Время финальное": created_time[maxi]}
        NotRec_Rec_Itog = NotRec_Rec_Itog.append(NotRec_Rec_row, ignore_index=True)

    elif status_ift[maxi] == 'Рекомендован' and status_ift[minim] == 'Рекомендован':
        Rec_Rec.append(release_key[maxi])
        Rec_Rec_row = {"Номер": release_key[minim], "Начальный статус": status_of_release[minim],
                          "Время начальное": created_time[minim], "Финальный статус": status_of_release[maxi],
                          "Время финальное": created_time[maxi]}
        Rec_Rec_Itog = Rec_Rec_Itog.append(Rec_Rec_row, ignore_index=True)

    elif status_ift[maxi] == 'Не рекомендован' and status_ift[minim] == 'Не рекомендован':
        NotRec_NotRec.append(release_key[maxi])
        NotRec_NotRec_row = {"Номер": release_key[minim], "Начальный статус": status_of_release[minim],
                          "Время начальное": created_time[minim], "Финальный статус": status_of_release[maxi],
                          "Время финальное": created_time[maxi]}
        NotRec_NotRec_Itog = NotRec_NotRec_Itog.append(NotRec_NotRec_row, ignore_index=True)

    elif status_ift[maxi] == 'Не рекомендован' and status_ift[minim] == 'Рекомендован':
        Rec_NotRec.append(release_key[maxi])
        Rec_NotRec_row = {"Номер": release_key[minim], "Начальный статус": status_of_release[minim],
                          "Время начальное": created_time[minim], "Финальный статус": status_of_release[maxi],
                          "Время финальное": created_time[maxi]}
        Rec_NotRec_Itog = Rec_NotRec_Itog.append(Rec_NotRec_row, ignore_index=True)

    else:
        trouble.append(release_key[maxi])

    new_rowmin = {"Id клика": ckick_id[minim], "Номер": release_key[minim],
                    "Время": created_time[minim], "Статус": status_of_release[minim],
                    "Статус": status_ift[minim]}
    new_rowmax = {"Id клика": ckick_id[maxi], "Номер": release_key[maxi], "Время": created_time[maxi],
                    "Статус": status_of_release[maxi], "Статус": status_ift[maxi]}

    return new_rowmin, new_rowmax


# Вызываем основной метод
for x in double1:
    rows = []
    method = relis1(x)
    rows = method
    df = df.append(rows[0], ignore_index=True)
    df = df.append(rows[1], ignore_index=True)


for i in range(len(trouble)):
    row = {"Номер": trouble[i]}
    trouble_df = trouble_df.append(row, ignore_index=True)

itog_df = pandas.DataFrame(
    {"Название массива": ['Всего', 'NotRec_NotRec', 'Rec_Rec', 'Rec_NotRec', 'NotRec_Rec', 'trouble'],
     "Количество": [len(double1), len(NotRec_NotRec), len(Rec_Rec), len(Rec_NotRec), len(NotRec_Rec),
                            len(trouble)],
     "Процент от общего числа": [' ', str(round((len(NotRec_NotRec) / len(double1) * 100), 2)) + "%",
                                         str(round((len(Rec_Rec) / len(double1) * 100), 2)) + "%",
                                         str(round((len(Rec_NotRec) / len(double1) * 100), 2)) + "%",
                                         str(round((len(NotRec_Rec) / len(double1) * 100), 2)) + "%",
                                         str(round((len(trouble) / len(double1) * 100), 2)) + "%"]
     })
# Сохраняем всё в эксель на нужные страницы
df.to_excel(writer, sheet_name='Лист итогов')
itog_df.to_excel(writer, sheet_name='Лист с процентами')

#NotRec_Rec_df.to_excel(writer, sheet_name='Не рекомендован->Рекомендован')
NotRec_Rec_Itog.to_excel(writer, sheet_name='Не рекомендован->Рекомендован')

#Rec_NotRec_df.to_excel(writer, sheet_name='Рекомендован->Не рекомендован')
Rec_NotRec_Itog.to_excel(writer, sheet_name='Рекомендован->Не рекомендован')

#Rec_Rec_df.to_excel(writer, sheet_name='Рекомендован->Рекомендован')
Rec_Rec_Itog.to_excel(writer, sheet_name='Рекомендован->Рекомендован')

#NotRec_NotRec_df.to_excel(writer, sheet_name='НеРекомендован->НеРекомендован')
NotRec_NotRec_Itog.to_excel(writer, sheet_name='НеРекомендован->НеРекомендован')

trouble_df.to_excel(writer, sheet_name='Error')

writer.save()

